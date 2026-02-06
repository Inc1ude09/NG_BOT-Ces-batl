import logging
import os
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    Update,
)
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

EXCEL_PATH = Path("case_battle_ledger.xlsx")
BOT_TOKEN = os.getenv("BOT_TOKEN", "")

MENU_KEYBOARD = ReplyKeyboardMarkup(
    [
        [KeyboardButton("💰 Пополнение"), KeyboardButton("💸 Вывод")],
        [KeyboardButton("💼 Баланс"), KeyboardButton("📊 Статистика")],
        [KeyboardButton("📝 История"), KeyboardButton("📤 Экспорт")],
        [KeyboardButton("🗑 Сброс")],
    ],
    resize_keyboard=True,
)


class LedgerStorage:
    """Хранилище транзакций в Excel (единый источник данных)."""

    TX_HEADERS = ["user_id", "type", "amount", "timestamp"]

    def __init__(self, file_path: Path):
        self.file_path = file_path
        self._init_workbook()

    def _init_workbook(self) -> None:
        if self.file_path.exists():
            return

        wb = openpyxl.Workbook()
        tx_sheet = wb.active
        tx_sheet.title = "Transactions"
        tx_sheet.append(self.TX_HEADERS)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col, header in enumerate(self.TX_HEADERS, 1):
            cell = tx_sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        summary = wb.create_sheet("Summary")
        summary.append(["user_id", "deposits", "withdrawals", "balance", "roi_percent", "updated_at"])
        for col in range(1, 7):
            cell = summary.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        wb.save(self.file_path)

    def _load_wb(self):
        return openpyxl.load_workbook(self.file_path)

    def _save_wb(self, wb) -> None:
        self._autosize_columns(wb["Transactions"])
        self._autosize_columns(wb["Summary"])
        wb.save(self.file_path)

    @staticmethod
    def _autosize_columns(sheet) -> None:
        for column in sheet.columns:
            max_length = 0
            for cell in column:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column[0].column_letter].width = max_length + 2

    def add_transaction(self, user_id: int, tx_type: str, amount: Decimal) -> None:
        wb = self._load_wb()
        tx = wb["Transactions"]
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        tx.append([user_id, tx_type, float(amount), timestamp])

        self._rebuild_summary(wb)
        self._save_wb(wb)

    def reset_user(self, user_id: int) -> None:
        wb = self._load_wb()
        tx = wb["Transactions"]
        kept_rows = [self.TX_HEADERS]

        for row in tx.iter_rows(min_row=2, values_only=True):
            if int(row[0]) != user_id:
                kept_rows.append(list(row))

        tx.delete_rows(1, tx.max_row)
        for row in kept_rows:
            tx.append(row)

        self._rebuild_summary(wb)
        self._save_wb(wb)

    def _rebuild_summary(self, wb) -> None:
        tx_sheet = wb["Transactions"]
        summary_sheet = wb["Summary"]
        summary_sheet.delete_rows(2, summary_sheet.max_row)

        user_stats: Dict[int, Dict[str, Decimal]] = {}
        for row in tx_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            uid = int(row[0])
            tx_type = row[1]
            amount = Decimal(str(row[2]))

            if uid not in user_stats:
                user_stats[uid] = {"deposit": Decimal("0"), "withdraw": Decimal("0")}
            user_stats[uid][tx_type] += amount

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for uid, stats in sorted(user_stats.items()):
            deposits = stats["deposit"]
            withdrawals = stats["withdraw"]
            balance = deposits - withdrawals
            roi = ((withdrawals - deposits) / deposits * Decimal("100")) if deposits > 0 else Decimal("0")
            summary_sheet.append([uid, float(deposits), float(withdrawals), float(balance), float(round(roi, 2)), now])

    def get_user_stats(self, user_id: int) -> Tuple[Decimal, Decimal, Decimal, Decimal]:
        wb = self._load_wb()
        summary = wb["Summary"]

        for row in summary.iter_rows(min_row=2, values_only=True):
            if row[0] and int(row[0]) == user_id:
                deposits = Decimal(str(row[1]))
                withdrawals = Decimal(str(row[2]))
                balance = Decimal(str(row[3]))
                roi = Decimal(str(row[4]))
                return deposits, withdrawals, balance, roi
        return Decimal("0"), Decimal("0"), Decimal("0"), Decimal("0")

    def get_user_history(self, user_id: int, limit: int = 10) -> List[Tuple[str, Decimal, str]]:
        wb = self._load_wb()
        tx = wb["Transactions"]

        rows = []
        for row in tx.iter_rows(min_row=2, values_only=True):
            if row[0] and int(row[0]) == user_id:
                rows.append((row[1], Decimal(str(row[2])), str(row[3])))
        return list(reversed(rows[-limit:]))


ledger = LedgerStorage(EXCEL_PATH)


def parse_amount(raw: str) -> Decimal:
    normalized = raw.replace(",", ".").strip()
    value = Decimal(normalized)
    if value <= 0:
        raise ValueError("amount must be positive")
    return value.quantize(Decimal("0.01"))


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    msg = (
        "🎮 <b>Case Battle Tracker</b>\n\n"
        "Используй всплывающее меню ниже или команды:\n"
        "• /add 1000\n"
        "• /withdraw 500\n"
        "• /balance\n"
        "• /stats\n"
        "• /history\n"
        "• /export\n"
        "• /reset"
    )
    await update.message.reply_text(msg, parse_mode="HTML", reply_markup=MENU_KEYBOARD)


async def add_deposit(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not context.args:
        await update.message.reply_text("❌ Укажи сумму: <code>/add 1000</code>", parse_mode="HTML")
        return
    try:
        amount = parse_amount(context.args[0])
    except (InvalidOperation, ValueError):
        await update.message.reply_text("❌ Неверная сумма. Пример: <code>/add 1000</code>", parse_mode="HTML")
        return

    user_id = update.effective_user.id
    ledger.add_transaction(user_id, "deposit", amount)
    _, _, balance, _ = ledger.get_user_stats(user_id)
    await update.message.reply_text(
        f"✅ Пополнение: <code>{amount:,.2f}</code> ₽\n💼 Баланс: <code>{balance:,.2f}</code> ₽",
        parse_mode="HTML",
    )


async def withdraw(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not context.args:
        await update.message.reply_text("❌ Укажи сумму: <code>/withdraw 500</code>", parse_mode="HTML")
        return
    try:
        amount = parse_amount(context.args[0])
    except (InvalidOperation, ValueError):
        await update.message.reply_text("❌ Неверная сумма. Пример: <code>/withdraw 500</code>", parse_mode="HTML")
        return

    user_id = update.effective_user.id
    ledger.add_transaction(user_id, "withdraw", amount)
    _, _, balance, _ = ledger.get_user_stats(user_id)
    await update.message.reply_text(
        f"✅ Вывод: <code>{amount:,.2f}</code> ₽\n💼 Баланс: <code>{balance:,.2f}</code> ₽",
        parse_mode="HTML",
    )


async def balance(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    _, _, user_balance, roi = ledger.get_user_stats(user_id)
    await update.message.reply_text(
        f"💼 Баланс: <code>{user_balance:,.2f}</code> ₽\n📈 ROI: <code>{roi:,.2f}%</code>",
        parse_mode="HTML",
    )


async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    deposits, withdrawals, user_balance, roi = ledger.get_user_stats(user_id)
    pnl = withdrawals - deposits
    await update.message.reply_text(
        "📊 <b>Статистика</b>\n\n"
        f"💰 Ввод: <code>{deposits:,.2f}</code> ₽\n"
        f"💸 Вывод: <code>{withdrawals:,.2f}</code> ₽\n"
        f"💼 Итого (баланс): <code>{user_balance:,.2f}</code> ₽\n"
        f"📈 ROI: <code>{roi:,.2f}%</code>\n"
        f"{'🎉 Прибыль' if pnl >= 0 else '💔 Убыток'}: <code>{abs(pnl):,.2f}</code> ₽",
        parse_mode="HTML",
    )


async def history(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    history_rows = ledger.get_user_history(user_id)
    if not history_rows:
        await update.message.reply_text("📝 История пуста.")
        return

    lines = ["📝 <b>Последние операции:</b>"]
    for tx_type, amount, timestamp in history_rows:
        title = "Пополнение" if tx_type == "deposit" else "Вывод"
        emoji = "💰" if tx_type == "deposit" else "💸"
        lines.append(f"{emoji} {title}: <code>{amount:,.2f}</code> ₽ — {timestamp}")
    await update.message.reply_text("\n".join(lines), parse_mode="HTML")


async def export_file(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not EXCEL_PATH.exists():
        await update.message.reply_text("❌ Файл еще не создан.")
        return
    with open(EXCEL_PATH, "rb") as file:
        await update.message.reply_document(
            document=file,
            filename=EXCEL_PATH.name,
            caption="📤 Выгрузка общей Excel-базы",
        )


async def reset(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    keyboard = [
        [
            InlineKeyboardButton("✅ Да, удалить", callback_data="reset_confirm"),
            InlineKeyboardButton("❌ Отмена", callback_data="reset_cancel"),
        ]
    ]
    await update.message.reply_text(
        "⚠️ Удалить все твои транзакции?",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )


async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    if query.data == "reset_confirm":
        ledger.reset_user(query.from_user.id)
        await query.edit_message_text("✅ Данные удалены")
    else:
        await query.edit_message_text("❌ Отмена")


async def menu_router(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    text = (update.message.text or "").strip()
    if text == "💰 Пополнение":
        context.user_data["awaiting_amount"] = "deposit"
        await update.message.reply_text("Введи сумму пополнения числом.", reply_markup=ReplyKeyboardRemove())
    elif text == "💸 Вывод":
        context.user_data["awaiting_amount"] = "withdraw"
        await update.message.reply_text("Введи сумму вывода числом.", reply_markup=ReplyKeyboardRemove())
    elif text == "💼 Баланс":
        await balance(update, context)
    elif text == "📊 Статистика":
        await stats(update, context)
    elif text == "📝 История":
        await history(update, context)
    elif text == "📤 Экспорт":
        await export_file(update, context)
    elif text == "🗑 Сброс":
        await reset(update, context)



async def amount_from_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    mode = context.user_data.get("awaiting_amount")
    if not mode:
        return

    try:
        amount = parse_amount(update.message.text)
    except (InvalidOperation, ValueError):
        await update.message.reply_text("❌ Введи корректное число.", reply_markup=MENU_KEYBOARD)
        context.user_data.pop("awaiting_amount", None)
        return

    context.user_data.pop("awaiting_amount", None)
    user_id = update.effective_user.id
    ledger.add_transaction(user_id, mode, amount)
    _, _, user_balance, _ = ledger.get_user_stats(user_id)
    action = "Пополнение" if mode == "deposit" else "Вывод"
    await update.message.reply_text(
        f"✅ {action}: <code>{amount:,.2f}</code> ₽\n💼 Баланс: <code>{user_balance:,.2f}</code> ₽",
        parse_mode="HTML",
        reply_markup=MENU_KEYBOARD,
    )


def main() -> None:
    if not BOT_TOKEN:
        raise RuntimeError("Set BOT_TOKEN env variable before run")

    app = Application.builder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("add", add_deposit))
    app.add_handler(CommandHandler("withdraw", withdraw))
    app.add_handler(CommandHandler("balance", balance))
    app.add_handler(CommandHandler("stats", stats))
    app.add_handler(CommandHandler("history", history))
    app.add_handler(CommandHandler("export", export_file))
    app.add_handler(CommandHandler("reset", reset))
    app.add_handler(CallbackQueryHandler(button_callback))
    app.add_handler(MessageHandler(filters.Regex(r"^(💰 Пополнение|💸 Вывод|💼 Баланс|📊 Статистика|📝 История|📤 Экспорт|🗑 Сброс)$"), menu_router))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, amount_from_menu))

    logger.info("Bot started")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
